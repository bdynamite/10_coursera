import random
import collections
import os

import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Alignment
import tqdm

course = collections.namedtuple('Course', 'name lang start weeks stars')


def get_courses_list():
    url = 'https://www.coursera.org/sitemap~www~courses.xml'
    quantity = 20
    responce = requests.get(url=url)
    root = ET.fromstring(responce.text)
    courses = [child[0].text for child in root]
    return random.sample(courses, quantity)


def get_course_info(course_slug):
    responce = requests.get(url=course_slug)
    soup = BeautifulSoup(responce.text, 'html.parser')
    course_name = soup.title.text.split('|')[0]
    lang = soup.find_all(attrs={'class': 'rc-Language'})[0].text
    start = soup.find_all(attrs={'class': 'startdate rc-StartDateString caption-text'})[0].text.split()[1:]
    start = ' '.join(start)
    weeks = len(soup.find_all(attrs={'class': 'week-body'}))
    try:
        stars = soup.find_all(attrs={'class': 'ratings-text headline-2-text'})[0].text.split()[1]
    except IndexError:
        stars = None
    return course(name=course_name, lang=lang, start=start, weeks=weeks, stars=stars)


def output_courses_info_to_xlsx(courses):
    filepath = 'coursera.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'courses'
    create_xlsx_header(ws)
    columns = 'ABCDE'
    for index, course in enumerate(courses):
        for elem_index, element in enumerate(course):
            cell = '{}{}'.format(columns[elem_index], index + 2)
            ws[cell] = element
            if cell[0] in 'AB':
                ws[cell].alignment = Alignment(wrapText=True)
    wb.save(filename=filepath)
    return filepath


def create_xlsx_header(sheet):
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['D'].width = 15
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Language'
    sheet['C1'] = 'Start date'
    sheet['D1'] = 'Duration (weeks)'
    sheet['E1'] = 'Rating'


def output_path(filepath):
    print('Courses saved at')
    print(os.path.join(os.getcwd(), filepath))
    

if __name__ == '__main__':
    courses = get_courses_list()
    courses_info = list(map(get_course_info, tqdm.tqdm(courses, desc='collecting data')))
    filepath = output_courses_info_to_xlsx(courses_info)
    output_path(filepath)
